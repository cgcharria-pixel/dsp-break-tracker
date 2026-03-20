"""
DSP Break Time Tracker – Analysis Engine
Generic / Template Version
"""

import re
import io
from datetime import datetime, date, timedelta

import pandas as pd
import numpy as np


# ─────────────────────────────────────────────
# ADP PARSING
# ─────────────────────────────────────────────

def _split_camelcase(name: str) -> str:
    """Insert space before each capital letter that follows a lowercase letter."""
    return re.sub(r"(?<=[a-z])(?=[A-Z])", " ", name)


def parse_adp(file) -> pd.DataFrame:
    """
    Parse an ADP weekly timecard export (.xlsx / .csv).

    Expected ADP columns (flexible detection):
        Employee Name | File # | Reg Hrs | OT Hrs | Total Hrs
        Punch-In Date/Time | Punch-Out Date/Time
    Returns a clean DataFrame with normalised columns.
    """
    try:
        if hasattr(file, "name") and file.name.endswith(".csv"):
            raw = pd.read_csv(file, header=None, dtype=str)
        else:
            raw = pd.read_excel(file, header=None, dtype=str)
    except Exception as e:
        raise ValueError(f"Cannot read ADP file: {e}")

    # ── find header row ────────────────────────────────────────────
    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v)]
        if any("employee" in v for v in vals):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Cannot locate header row in ADP file (looking for 'Employee').")

    df = raw.iloc[header_row + 1:].copy()
    df.columns = [str(c).strip() for c in raw.iloc[header_row]]
    df = df.dropna(how="all").reset_index(drop=True)

    # ── column detection ───────────────────────────────────────────
    col_map = {}
    for col in df.columns:
        lc = col.lower()
        if "employee" in lc and "name" in lc:
            col_map["employee_name"] = col
        elif "file" in lc and "#" in lc:
            col_map["file_number"] = col
        elif "reg" in lc and "hr" in lc:
            col_map["reg_hrs"] = col
        elif "ot" in lc and "hr" in lc:
            col_map["ot_hrs"] = col
        elif "total" in lc and "hr" in lc:
            col_map["total_hrs"] = col
        elif "punch" in lc and "in" in lc:
            col_map["punch_in"] = col
        elif "punch" in lc and "out" in lc:
            col_map["punch_out"] = col

    # ── build output ───────────────────────────────────────────
    result = pd.DataFrame()
    result["employee_name"] = (
        df[col_map["employee_name"]].astype(str).str.strip()
        if "employee_name" in col_map else "Unknown"
    )
    result["file_number"] = (
        df[col_map["file_number"]].astype(str).str.strip()
        if "file_number" in col_map else ""
    )

    for field in ("reg_hrs", "ot_hrs", "total_hrs"):
        if field in col_map:
            result[field] = pd.to_numeric(df[col_map[field]], errors="coerce").fillna(0.0)
        else:
            result[field] = 0.0

    for field in ("punch_in", "punch_out"):
        if field in col_map:
            result[field] = pd.to_datetime(df[col_map[field]], errors="coerce")
        else:
            result[field] = pd.NaT

    # ── compute ADP shift duration ───────────────────────────────────────
    if result["punch_in"].notna().any() and result["punch_out"].notna().any():
        result["adp_duration_min"] = (
            (result["punch_out"] - result["punch_in"]).dt.total_seconds() / 60
        ).round(1)
    else:
        result["adp_duration_min"] = result["total_hrs"] * 60

    # ── clean employee names ──────────────────────────────────────────
    result = result[result["employee_name"].str.len() > 1].copy()
    result["employee_name"] = result["employee_name"].apply(
        lambda n: _split_camelcase(n) if n == n.title() else n
    )

    return result.reset_index(drop=True)


# ─────────────────────────────────────────────
# AMAZON PARSING
# ─────────────────────────────────────────────

def parse_amazon(file) -> pd.DataFrame:
    """
    Parse an Amazon Flex / DSP time export (.xlsx / .csv).

    Expected columns (flexible):
        Associate Name | ID / Badge | Shift Start | Shift End | Duration
    Returns a clean DataFrame.
    """
    try:
        if hasattr(file, "name") and file.name.endswith(".csv"):
            raw = pd.read_csv(file, header=None, dtype=str)
        else:
            raw = pd.read_excel(file, header=None, dtype=str)
    except Exception as e:
        raise ValueError(f"Cannot read Amazon file: {e}")

    # ── find header row ────────────────────────────────────────────
    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v)]
        if any("associate" in v or "employee" in v or "name" in v for v in vals):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Cannot locate header row in Amazon file (looking for 'Associate' or 'Name').")

    df = raw.iloc[header_row + 1:].copy()
    df.columns = [str(c).strip() for c in raw.iloc[header_row]]
    df = df.dropna(how="all").reset_index(drop=True)

    # ── column detection ───────────────────────────────────────────
    col_map = {}
    for col in df.columns:
        lc = col.lower()
        if ("associate" in lc or "employee" in lc) and "name" in lc:
            col_map["associate_name"] = col
        elif "name" in lc and "associate" not in col_map:
            col_map["associate_name"] = col
        elif "id" in lc or "badge" in lc:
            col_map["associate_id"] = col
        elif "start" in lc:
            col_map["shift_start"] = col
        elif "end" in lc:
            col_map["shift_end"] = col
        elif "duration" in lc or "hours" in lc:
            col_map["duration"] = col

    # ── build output ───────────────────────────────────────────
    result = pd.DataFrame()
    result["associate_name"] = (
        df[col_map["associate_name"]].astype(str).str.strip()
        if "associate_name" in col_map else "Unknown"
    )
    result["associate_id"] = (
        df[col_map["associate_id"]].astype(str).str.strip()
        if "associate_id" in col_map else ""
    )

    for field in ("shift_start", "shift_end"):
        if field in col_map:
            result[field] = pd.to_datetime(df[col_map[field]], errors="coerce")
        else:
            result[field] = pd.NaT

    # ── compute Amazon shift duration ───────────────────────────────────────
    if result["shift_start"].notna().any() and result["shift_end"].notna().any():
        result["amz_duration_min"] = (
            (result["shift_end"] - result["shift_start"]).dt.total_seconds() / 60
        ).round(1)
    elif "duration" in col_map:
        result["amz_duration_min"] = pd.to_numeric(df[col_map["duration"]], errors="coerce").fillna(0.0) * 60
    else:
        result["amz_duration_min"] = 0.0

    result = result[result["associate_name"].str.len() > 1].copy()

    return result.reset_index(drop=True)


# ─────────────────────────────────────────────
# MATCHING
# ─────────────────────────────────────────────

def _normalize_name(name: str) -> str:
    """Lowercase, strip punctuation/spaces for fuzzy matching."""
    name = str(name).lower().strip()
    name = re.sub(r"[^a-z\s]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name


def _last_first(name: str) -> str:
    """Convert 'First Last' → 'last first' for matching."""
    parts = name.strip().split()
    if len(parts) >= 2:
        return f"{parts[-1]} {' '.join(parts[:-1])}"
    return name


def match_employees(adp_df: pd.DataFrame, amazon_df: pd.DataFrame) -> pd.DataFrame:
    """
    Fuzzy-match ADP employees to Amazon associates by name.
    Returns merged DataFrame with both sets of columns.
    """
    adp = adp_df.copy()
    amz = amazon_df.copy()

    adp["_key"] = adp["employee_name"].apply(_normalize_name)
    adp["_key_lf"] = adp["_key"].apply(_last_first)

    amz["_key"] = amz["associate_name"].apply(_normalize_name)
    amz["_key_lf"] = amz["_key"].apply(_last_first)

    # ── exact match on normalised name ──────────────────────────────────────
    merged = adp.merge(amz, on="_key", how="outer", suffixes=("_adp", "_amz"))

    # ── second pass: last-first swap for unmatched ───────────────────────────
    unmatched_adp = merged[merged["associate_name"].isna()]["employee_name"].tolist()
    unmatched_amz = merged[merged["employee_name"].isna()]["associate_name"].tolist()

    if unmatched_adp and unmatched_amz:
        adp2 = adp[adp["employee_name"].isin(unmatched_adp)].copy()
        amz2 = amz[amz["associate_name"].isin(unmatched_amz)].copy()
        merged2 = adp2.merge(amz2, left_on="_key_lf", right_on="_key", how="inner", suffixes=("_adp", "_amz"))
        if not merged2.empty:
            merged = pd.concat([merged[~merged["employee_name"].isin(unmatched_adp)], merged2], ignore_index=True)

    merged.drop(columns=[c for c in merged.columns if c.startswith("_")], inplace=True, errors="ignore")
    return merged.reset_index(drop=True)


# ─────────────────────────────────────────────
# DISCREPANCY CALCULATION
# ─────────────────────────────────────────────

BREAK_THRESHOLDS = {
    "short":  {"min_hrs": 0,  "max_hrs": 6,  "break_min": 0},
    "medium": {"min_hrs": 6,  "max_hrs": 8,  "break_min": 30},
    "long":   {"min_hrs": 8,  "max_hrs": 10, "break_min": 30},
    "extended": {"min_hrs": 10, "max_hrs": 99, "break_min": 60},
}


def _required_break(adp_min: float) -> int:
    """Return required unpaid break in minutes based on ADP shift length."""
    hrs = adp_min / 60
    if hrs < 6:
        return 0
    elif hrs < 8:
        return 30
    elif hrs < 10:
        return 30
    else:
        return 60


def calculate_discrepancies(df: pd.DataFrame) -> pd.DataFrame:
    """
    Given a matched DataFrame, compute:
      - expected Amazon duration (ADP – required break)
      - discrepancy (Amazon – expected)
      - discrepancy flag
    """
    result = df.copy()

    # coerce
    result["adp_duration_min"] = pd.to_numeric(result.get("adp_duration_min", 0), errors="coerce").fillna(0)
    result["amz_duration_min"] = pd.to_numeric(result.get("amz_duration_min", 0), errors="coerce").fillna(0)

    result["required_break_min"] = result["adp_duration_min"].apply(_required_break)
    result["expected_amz_min"]   = result["adp_duration_min"] - result["required_break_min"]
    result["discrepancy_min"]    = (result["amz_duration_min"] - result["expected_amz_min"]).round(1)

    def classify(d):
        if abs(d) <= 5:
            return "✅ OK"
        elif d > 5:
            return "⚠️ Over"
        else:
            return "🔴 Under"

    result["status"] = result["discrepancy_min"].apply(classify)

    # human-readable hours
    for col in ("adp_duration_min", "amz_duration_min", "expected_amz_min", "discrepancy_min"):
        h_col = col.replace("_min", "_hrs")
        result[h_col] = (result[col] / 60).round(2)

    return result.reset_index(drop=True)


# ─────────────────────────────────────────────
# SCRIPT BUILDER
# ─────────────────────────────────────────────

def build_script(row: pd.Series) -> str:
    """
    Build a call / conversation script for a single discrepancy row.
    """
    name    = row.get("employee_name", row.get("associate_name", "Associate"))
    disc    = row.get("discrepancy_min", 0)
    adp_h   = row.get("adp_duration_hrs", 0)
    amz_h   = row.get("amz_duration_hrs", 0)
    brk_min = int(row.get("required_break_min", 0))
    status  = row.get("status", "")

    if "OK" in str(status):
        return (
            f"Hi {name}, I'm reviewing this week's time records and everything looks good on your end. "
            f"ADP shows {adp_h:.2f} hrs and Amazon shows {amz_h:.2f} hrs — right on target. "
            f"Thank you for keeping your records accurate!"
        )

    direction = "over" if disc > 0 else "under"
    diff_min  = abs(round(disc))
    diff_h    = abs(disc / 60)

    opener = f"Hi {name}, I'm reviewing this week's time records and I wanted to discuss a discrepancy I found."

    if direction == "over":
        body = (
            f" ADP shows your shift as {adp_h:.2f} hrs. With your {brk_min}-minute required unpaid break, "
            f"Amazon should show {(adp_h - brk_min/60):.2f} hrs. "
            f"However, Amazon is showing {amz_h:.2f} hrs — that's {diff_min} minutes ({diff_h:.2f} hrs) MORE than expected. "
            f"This could indicate the break wasn't logged correctly or the punch-out time needs to be reviewed."
        )
    else:
        body = (
            f" ADP shows your shift as {adp_h:.2f} hrs. With your {brk_min}-minute required unpaid break, "
            f"Amazon should show {(adp_h - brk_min/60):.2f} hrs. "
            f"However, Amazon is showing {amz_h:.2f} hrs — that's {diff_min} minutes ({diff_h:.2f} hrs) LESS than expected. "
            f"This could mean a punch-in/out was missed or the break was recorded as longer than it was."
        )

    action = (
        " Can you please review your records and let me know if any corrections are needed? "
        "We want to make sure everything is accurate before the week closes."
    )

    return opener + body + action


# ─────────────────────────────────────────────
# FULL PIPELINE
# ─────────────────────────────────────────────

def detect_station(adp_name: str, amz_name: str) -> str:
    """
    Try to infer the station code from filenames.
    Returns the station code or '' if not found.
    """
    pattern = r"\b([DHJWX][A-Z]{2}\d{1,2}|JFK\d|LGB\d|MKE\d)\b"
    for name in (adp_name or "", amz_name or ""):
        m = re.search(pattern, name.upper())
        if m:
            return m.group(1)
    return ""


def run_analysis(adp_file, amazon_file) -> pd.DataFrame:
    """
    End-to-end: parse both files → match → calculate discrepancies.
    Returns the final enriched DataFrame.
    """
    adp_df = parse_adp(adp_file)
    amz_df = parse_amazon(amazon_file)
    merged = match_employees(adp_df, amz_df)
    result = calculate_discrepancies(merged)
    return result


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def export_excel(df: pd.DataFrame, report_date: date, station: str) -> bytes:
    """
    Export the discrepancy report as a styled Excel workbook.
    Returns bytes suitable for st.download_button.
    """
    output = io.BytesIO()

    # ── select / rename display columns ───────────────────────────────────────
    display_cols = {
        "employee_name":      "Employee Name",
        "file_number":        "File #",
        "adp_duration_hrs":   "ADP Hours",
        "amz_duration_hrs":   "Amazon Hours",
        "required_break_min": "Required Break (min)",
        "expected_amz_hrs":   "Expected Amazon Hrs",
        "discrepancy_hrs":    "Discrepancy (hrs)",
        "discrepancy_min":    "Discrepancy (min)",
        "status":             "Status",
    }
    available = {k: v for k, v in display_cols.items() if k in df.columns}
    export_df = df[list(available.keys())].rename(columns=available)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Break Discrepancies")

        ws = writer.sheets["Break Discrepancies"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header style
        header_fill  = PatternFill("solid", fgColor="1B3A6B")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border

        # Data rows
        ok_fill   = PatternFill("solid", fgColor="D6F5E3")
        over_fill = PatternFill("solid", fgColor="FFF3CD")
        bad_fill  = PatternFill("solid", fgColor="FDDEDE")

        status_col = None
        for idx, col in enumerate(export_df.columns, 1):
            if col == "Status":
                status_col = idx
                break

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="center")
            if status_col:
                status_val = str(row[status_col - 1].value or "")
                if "OK" in status_val:
                    fill = ok_fill
                elif "Over" in status_val:
                    fill = over_fill
                else:
                    fill = bad_fill
                for cell in row:
                    cell.fill = fill

        # Auto-fit columns
        for col_idx, col in enumerate(export_df.columns, 1):
            max_len = max(len(str(col)), *(len(str(v)) for v in export_df.iloc[:, col_idx - 1]))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Add metadata sheet
        meta_ws = writer.book.create_sheet("Report Info")
        meta_ws["A1"] = "Report Date"
        meta_ws["B1"] = str(report_date)
        meta_ws["A2"] = "Station"
        meta_ws["B2"] = station
        meta_ws["A3"] = "Generated"
        meta_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        meta_ws["A4"] = "Total Records"
        meta_ws["B4"] = len(export_df)
        meta_ws["A5"] = "Discrepancies Found"
        meta_ws["B5"] = len(export_df[~export_df["Status"].str.contains("OK", na=False)])

    return output.getvalue()
